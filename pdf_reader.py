import PyPDF2 
import textract

import nltk
nltk.download('punkt')
nltk.download('stopwords')

from nltk.tokenize import word_tokenize
from nltk.corpus import stopwords

import collections
import pycountry
import re
import string

from openpyxl import load_workbook
import os, errno, sys
import shutil
import glob

# country checklist
country_list = {}
for country in pycountry.countries:
    country_list[country.name.lower()] = country.name
    country_list[country.alpha_3.lower()] = country.name

# candidature checklist
candidature_list = {
    'master':'Master',
    'masters':'Master',
    'bachelor':'Bachelor',
    'dr':'PhD',
    'phd':'PhD',
    'doctor':'PhD',
    'doctorate':'PhD',
    'm.s.':'Master',
    'b.s.':'Bachelor',
    'ms':'Master'
}

# grant keywords
travelgrant_list = ['grant','scholarship','financial','aids','aid']

# process text
def text_preprocessing(text):
    # remove numbers
    text = re.sub(r'\d+','',text)
    # remove symbols
    text = text.translate(str.maketrans('','', string.punctuation))
    text = text.strip()
    tokens = word_tokenize(text)
    stop_words = stopwords.words('english')
    keywords = [word.lower() for word in tokens if not word in stop_words]
    keywords_counter = collections.Counter(keywords)

    return keywords_counter


# --------------------------
# -------- Main ------------
# --------------------------

ref_dict = {}
ref_list = []

# load reference set and load to dictionary
ref_fpath= 'summerschool_ref.xlsx'
ref_book = load_workbook(ref_fpath)
ref_namelist = ref_book.get_sheet_names()
ref_meta_sheet = ref_book.get_sheet_by_name(ref_namelist[0])

for row_idx in range(ref_meta_sheet.min_row+1, ref_meta_sheet.max_row+1):

    if ref_meta_sheet.cell(row=row_idx, column=6).value == 'Student':

        # get id
        id_num_str = ref_meta_sheet.cell(row=row_idx, column=1).value
        id_num = id_num_str.replace('1-','')

        # get names
        fullname = ''
        firstname = ref_meta_sheet.cell(row=row_idx, column=2).value
        lastname = ref_meta_sheet.cell(row=row_idx, column=3).value
        if firstname:
            fullname = fullname + firstname
        if lastname:
            fullname += ' '
            fullname += lastname

        # get email
        email = ref_meta_sheet.cell(row=row_idx, column=4).value
        if not email:
            email = '*NIL*'
        
        affiliation = ref_meta_sheet.cell(row=row_idx, column=5).value
        if not affiliation:
            affiliation = '*NIL*'

        ref_dict[id_num] = [fullname, email, affiliation]


target_dict = {}


# load residual files
path = 'Sorted/Student/'
#files = [f for f in glob.glob(path + "**/*.pdf", recursive=True)]
files = [name for name in os.listdir(path)]

for fpath in files:
    _numid = fpath.split('_')[0]
    if _numid in ref_dict.keys():

        # collect meta infor
        target_dict[_numid] = ref_dict[_numid]

        country = '*NIL*'
        candidature = '*NIL*'
        need_grant = 'NO'

        # check if cv exist
        path_to_cv = path + fpath + '/cv.pdf'
        path_to_lom = path + fpath + '/letter_of_motivation.pdf'

        if os.path.exists(path_to_cv):
            cv = textract.process(path_to_cv, method='tesseract', language='eng')
            cv = cv.decode('utf-8')
            cv_keywords = text_preprocessing(cv)

            for key in cv_keywords.keys():
                if key in country_list.keys():
                    country = country_list[key]
                    break

            for key in cv_keywords.keys():
                if key in candidature_list.keys():
                    candidature = candidature_list[key]
                    break

        if os.path.exists(path_to_lom):
            lom = textract.process(path_to_lom, method='tesseract', language='eng')
            lom = lom.decode('utf-8')
            lom_keywords = text_preprocessing(lom)

            for key in lom_keywords.keys():
                if key in travelgrant_list:
                    need_grant = 'YES'
                    break

        target_dict[_numid].append(country)
        target_dict[_numid].append(candidature)
        target_dict[_numid].append(need_grant)


# load writing to workbook
tar_fpath= 'summerschool_ver_c.xlsx'
tar_book = load_workbook(tar_fpath)
tar_namelist = tar_book.get_sheet_names()
tar_meta_sheet = tar_book.get_sheet_by_name(tar_namelist[1])

for index in range(67, len(target_dict)):
    tar_meta_sheet.cell(row=index, column=1).value = '0'

tar_book.save('summerschool_test.xlsx')




